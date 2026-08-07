import telebot
from telebot import types
import openpyxl
import json
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

# BotFather'dan olgan tokeningiz (Railway'ning Variables bo'limidan olinadi)
BOT_TOKEN = os.getenv("BOT_TOKEN")

# Baza.xlsx faylining nomi
EXCEL_FILE = "Baza.xlsx"

bot = telebot.TeleBot(BOT_TOKEN)

# Admin ID
ADMIN_ID = 199728470

# "ZAKAZ NORTIX" guruhi ID raqami (sklad shu yerda buyurtmalarni qabul qiladi)
# Guruhga botni admin qilib qo'shing, guruh ichida /groupid deb yozing — bot sizga ID'ni yuboradi.
# Keyin shu qatordagi 0 o'rniga o'sha ID'ni (masalan -1001234567890) yozing.
ZAKAZ_GRUPPA_ID = -5166542981

# Holat xotiralari
yangilash_holati = {}
rasm_kutilayotganlar = {}
aksiya_kutilayotganlar = {}
buyurtma_holati = {}
qidiruv_kutayotganlar = set()
savat = {}
kutilayotgan_buyurtmalar = {}
buyurtma_id_hisoblagich = {"son": 0}


def ombor_malumotlarini_oqish():
    workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook.active

    malumotlar = {}
    for row in sheet.iter_rows(min_row=1, values_only=True):
        # Ustunlar tartibi: A=Brendlar, B=Kategoriya, C=Model, D=SONI, E=Narx, F=Rasm, G=Aksiya
        brend = row[0]
        kategoriya, model, soni = row[1], row[2], row[3]
        narxi = row[4] if len(row) > 4 else None
        rasm = row[5] if len(row) > 5 else None
        aksiya = row[6] if len(row) > 6 else None

        if kategoriya is None or model is None:
            continue
        kategoriya = str(kategoriya).strip()
        model = str(model).strip()
        brend = str(brend).strip() if brend is not None else ""
        aksiya = str(aksiya).strip() if aksiya is not None else ""

        if kategoriya.lower() == "kategoriya" or model.lower() == "model":
            continue

        soni = soni if soni is not None else 0
        narxi = narxi if narxi is not None else 0

        if kategoriya not in malumotlar:
            malumotlar[kategoriya] = []
        malumotlar[kategoriya].append((model, soni, narxi, rasm, brend, aksiya))

    return malumotlar


def ombor_sonini_yangilash(kategoriya, model, yangi_soni):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1):
        row_kategoriya = row[1].value
        row_model = row[2].value
        if row_kategoriya is None or row_model is None:
            continue
        if str(row_kategoriya).strip() == kategoriya and str(row_model).strip() == model:
            row[3].value = yangi_soni
            workbook.save(EXCEL_FILE)
            return True

    return False


def ombor_rasmini_yangilash(kategoriya, model, rasm_manzili):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1):
        row_kategoriya = row[1].value
        row_model = row[2].value
        if row_kategoriya is None or row_model is None:
            continue
        if str(row_kategoriya).strip() == kategoriya and str(row_model).strip() == model:
            if len(row) > 5:
                row[5].value = rasm_manzili
                workbook.save(EXCEL_FILE)
                return True
            return False

    return False


def ombor_aksiyasini_yangilash(kategoriya, model, aksiya_matni):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1):
        row_kategoriya = row[1].value
        row_model = row[2].value
        if row_kategoriya is None or row_model is None:
            continue
        if str(row_kategoriya).strip() == kategoriya and str(row_model).strip() == model:
            sheet.cell(row=row[0].row, column=7).value = aksiya_matni
            workbook.save(EXCEL_FILE)
            return True

    return False


BUYURTMALAR_FAYLI = "buyurtmalar.json"


def buyurtma_raqami(buyurtma_id):
    return f"INL-{buyurtma_id:07d}"


def buyurtma_pdf_yaratish(buyurtma_id, buyurtma):
    """Buyurtma uchun PDF fayl (invoys) yaratadi va fayl yo'lini qaytaradi."""
    fayl_nomi = f"/tmp/{buyurtma_raqami(buyurtma_id)}.pdf"
    c = canvas.Canvas(fayl_nomi, pagesize=A4)
    width, height = A4
    y = height - 60

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, f"Buyurtma {buyurtma_raqami(buyurtma_id)}")
    y -= 25

    c.setFont("Helvetica", 11)
    c.drawString(50, y, f"Sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    y -= 18
    c.drawString(50, y, f"Mijoz: {buyurtma['ism']}")
    y -= 18
    c.drawString(50, y, f"Telefon: {buyurtma['telefon']}")
    y -= 18
    c.drawString(50, y, f"Manzil: {buyurtma['manzil']}")
    y -= 30

    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Mahsulotlar:")
    y -= 20

    c.setFont("Helvetica", 10)
    for item in buyurtma["itemlar"]:
        summa = item["son"] * item["narxi"]
        qator = f"{item['model']} ({item['kategoriya']}) — {item['son']} dona x ${item['narxi']:,.2f} = ${summa:,.2f}"
        c.drawString(60, y, qator)
        y -= 16
        if y < 60:
            c.showPage()
            y = height - 60
            c.setFont("Helvetica", 10)

    y -= 15
    c.setFont("Helvetica-Bold", 13)
    c.drawString(50, y, f"Jami: ${buyurtma['jami_summa']:,.2f}")

    c.save()
    return fayl_nomi


def buyurtmani_saqlash(chat_id, itemlar, ism, telefon, manzil, jami_summa, holat="tasdiqlangan", buyurtma_id=None):
    try:
        if os.path.exists(BUYURTMALAR_FAYLI):
            with open(BUYURTMALAR_FAYLI, "r", encoding="utf-8") as f:
                barcha = json.load(f)
        else:
            barcha = []
    except Exception:
        barcha = []

    barcha.append({
        "buyurtma_id": buyurtma_id,
        "chat_id": chat_id,
        "sana": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "itemlar": itemlar,
        "ism": ism,
        "telefon": telefon,
        "manzil": manzil,
        "jami_summa": jami_summa,
        "holat": holat,
    })

    with open(BUYURTMALAR_FAYLI, "w", encoding="utf-8") as f:
        json.dump(barcha, f, ensure_ascii=False, indent=2)


def mijoz_buyurtmalarini_olish(chat_id):
    try:
        with open(BUYURTMALAR_FAYLI, "r", encoding="utf-8") as f:
            barcha = json.load(f)
    except Exception:
        return []

    mijoznikilar = [b for b in barcha if b.get("chat_id") == chat_id]
    return list(reversed(mijoznikilar))


def barcha_buyurtmalarni_indeks_bilan_olish():
    try:
        with open(BUYURTMALAR_FAYLI, "r", encoding="utf-8") as f:
            barcha = json.load(f)
    except Exception:
        return []

    return list(reversed(list(enumerate(barcha))))


KONTAKTLAR = [
    ("Muhiddin", "+998 91 999 40 30"),
    ("Durdivoy", "+998 91 422 27 72"),
    ("Murod", "+998 90 436 63 63"),
    ("Zafar", "+998 97 033 39 39"),
]
KONTAKT_MANZIL = "Toshkent city, Boulevar"
ISH_VAQTI = (
    "Dushanba: 9:30-18:30\n"
    "Seshanba: 9:30-18:30\n"
    "Chorshanba: 9:30-18:30\n"
    "Payshanba: 9:30-18:30\n"
    "Juma: 9:30-18:30\n"
    "Shanba: 9:30-18:30\n"
    "Yakshanba: Dam olish kuni"
)


def bosh_menyu_yaratish(user_id=None):
    """Asosiy pastki menyu tugmalari (Brend qo'shildi)"""
    menyu = types.ReplyKeyboardMarkup(resize_keyboard=True)
    menyu.add(types.KeyboardButton("🏷 Brendlar"), types.KeyboardButton("🔍 Qidiruv"))
    menyu.add(types.KeyboardButton("🛒 Savatim"), types.KeyboardButton("🧾 Buyurtmalarim"))
    menyu.add(types.KeyboardButton("☎️ Biz bilan bog'lanish"), types.KeyboardButton("📍 Manzilimiz"))
    menyu.add(types.KeyboardButton("ℹ️ Bot haqida"))

    if str(user_id) == str(ADMIN_ID):
        menyu.add(types.KeyboardButton("📷 Mahsulot rasmi"), types.KeyboardButton("🎉 Aksiya qo'shish"))

    return menyu


def orqaga_menyu_yaratish():
    menyu = types.ReplyKeyboardMarkup(resize_keyboard=True)
    menyu.add(types.KeyboardButton("🔙 Orqaga"), types.KeyboardButton("🏠 Bosh menyu"))
    return menyu


@bot.message_handler(func=lambda message: message.text in ["🔙 Orqaga", "🏠 Bosh menyu"])
def cancel_and_back(message):
    chat_id = message.chat.id
    user_id = message.from_user.id

    buyurtma_holati.pop(chat_id, None)
    yangilash_holati.pop(user_id, None)
    rasm_kutilayotganlar.pop(user_id, None)
    aksiya_kutilayotganlar.pop(user_id, None)
    qidiruv_kutayotganlar.discard(chat_id)

    bot.send_message(
        chat_id,
        "🏠 Asosiy menyuga qaytdingiz.",
        reply_markup=bosh_menyu_yaratish(user_id)
    )


@bot.message_handler(commands=["groupid"])
def groupid_handler(message):
    bot.send_message(message.chat.id, f"🆔 Ushbu chat ID: `{message.chat.id}`", parse_mode="Markdown")


@bot.message_handler(commands=['start'])
def start_handler(message):
    bot.send_message(
        message.chat.id,
        "Assalomu alaykum! Xorazm baza savdo botiga xush kelibsiz.\n\n"
        "Bu bot orqali siz ombordagi mahsulotlar qoldig'ini ko'rishingiz "
        "va zakaz berishingiz mumkin bo'ladi.",
        reply_markup=bosh_menyu_yaratish(message.from_user.id)
    )
    kategoriyalarni_korsatish(message.chat.id)


# ==========================================
# BRENDLAR BO'LIMI 
# ==========================================
@bot.message_handler(func=lambda message: message.text == "🏷 Brendlar")
def menyu_brendlar(message):
    keyboard = types.InlineKeyboardMarkup(row_width=2)
    brendlar = ["SAMSUNG", "PREMIER", "SONOR", "AUFIT"]
    
    tugmalar = [types.InlineKeyboardButton(text=f"🏢 {brend}", callback_data=f"brend:{brend}") for brend in brendlar]
    keyboard.add(*tugmalar)
        
    bot.send_message(
        message.chat.id, 
        "Qaysi brenddagi mahsulotlarni ko'rmoqchisiz?", 
        reply_markup=keyboard
    )
    bot.send_message(
        message.chat.id, 
        "Ortga qaytish uchun pastdagi '🔙 Orqaga' tugmasini bosing:",
        reply_markup=orqaga_menyu_yaratish()
    )


@bot.callback_query_handler(func=lambda call: call.data.startswith("brend:"))
def brend_tanlandi(call):
    bot.answer_callback_query(call.id)
    tanlangan_brend = call.data.split("brend:", 1)[1]

    try:
        malumotlar = ombor_malumotlarini_oqish()
    except Exception as e:
        bot.send_message(call.message.chat.id, f"Xatolik yuz berdi: {e}")
        return

    kategoriyalar = []
    for kategoriya, mahsulotlar in malumotlar.items():
        for item in mahsulotlar:
            soni = item[1]
            brend = item[4] if len(item) > 4 else ""
            if brend.lower() == tanlangan_brend.lower() and soni and soni > 0:
                if kategoriya not in kategoriyalar:
                    kategoriyalar.append(kategoriya)
                break

    if not kategoriyalar:
        bot.send_message(
            call.message.chat.id,
            f"Kechirasiz, <b>{tanlangan_brend}</b> brendiga oid mahsulotlar hozircha omborda yo'q.",
            parse_mode="HTML"
        )
        return

    keyboard = types.InlineKeyboardMarkup()
    for kategoriya in kategoriyalar:
        keyboard.add(types.InlineKeyboardButton(
            text=kategoriya,
            callback_data=f"brendkat:{tanlangan_brend}|{kategoriya}"
        ))

    bot.send_message(
        call.message.chat.id,
        f"🏢 <b>{tanlangan_brend}</b> — qaysi kategoriyani ko'rmoqchisiz?",
        parse_mode="HTML",
        reply_markup=keyboard
    )


@bot.callback_query_handler(func=lambda call: call.data.startswith("brendkat:"))
def brend_kategoriya_tanlandi(call):
    bot.answer_callback_query(call.id)
    tanlangan_brend, kategoriya = call.data.split("brendkat:", 1)[1].split("|", 1)

    try:
        malumotlar = ombor_malumotlarini_oqish()
    except Exception as e:
        bot.send_message(call.message.chat.id, f"Xatolik yuz berdi: {e}")
        return

    mahsulotlar = malumotlar.get(kategoriya, [])
    topilganlar = []
    for item in mahsulotlar:
        model, soni = item[0], item[1]
        narxi = item[2] if len(item) > 2 and item[2] else 0
        brend = item[4] if len(item) > 4 else ""
        aksiya_bor = "🎉 " if (len(item) > 5 and item[5]) else ""
        if brend.lower() == tanlangan_brend.lower() and soni and soni > 0:
            topilganlar.append((model, soni, narxi, aksiya_bor))

    if not topilganlar:
        bot.send_message(call.message.chat.id, "Bu kategoriyada hozircha omborda mahsulot yo'q.")
        return

    keyboard = types.InlineKeyboardMarkup()
    for model, soni, narxi, aksiya_bor in topilganlar:
        keyboard.add(types.InlineKeyboardButton(
            text=f"{aksiya_bor}{model} — ${narxi:,.2f}",
            callback_data=f"buy:{kategoriya}|{model}"
        ))

    bot.send_message(
        call.message.chat.id,
        f"🏢 <b>{tanlangan_brend}</b> — <b>{kategoriya}</b>:",
        parse_mode="HTML",
        reply_markup=keyboard
    )
# ==========================================


# ADMIN RASM BO'LIMI FUNKSIYASI
def mahsulot_rasmi_menyu_chiqarish(chat_id, user_id):
    if str(user_id) != str(ADMIN_ID):
        return

    try:
        malumotlar = ombor_malumotlarini_oqish()
    except Exception as e:
        bot.send_message(chat_id, f"Xatolik yuz berdi: {e}")
        return

    if not malumotlar:
        bot.send_message(chat_id, "Omborda mahsulot yo'q.")
        return

    keyboard = types.InlineKeyboardMarkup()
    for kategoriya in malumotlar.keys():
        keyboard.add(types.InlineKeyboardButton(
            text=f"📁 {kategoriya}",
            callback_data=f"imgkat:{kategoriya}"
        ))

    keyboard.add(types.InlineKeyboardButton("🔙 Orqaga", callback_data="back_to_main_admin"))

    bot.send_message(
        chat_id,
        "📷 <b>Rasm qo'shish bo'limi:</b>\nQaysi kategoriyadagi mahsulotga rasm biriktirasiz?",
        parse_mode="HTML",
        reply_markup=keyboard
    )
    bot.send_message(
        chat_id,
        "Ortga qaytish uchun '🔙 Orqaga' tugmasini bosing:",
        reply_markup=orqaga_menyu_yaratish()
    )


@bot.message_handler(func=lambda message: message.text == "📷 Mahsulot rasmi")
def mahsulot_rasmi_boshlash(message):
    mahsulot_rasmi_menyu_chiqarish(message.chat.id, message.from_user.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("imgkat:"))
def rasm_kategoriya_tanlandi(call):
    bot.answer_callback_query(call.id)
    if str(call.from_user.id) != str(ADMIN_ID):
        return

    kategoriya = call.data.split("imgkat:", 1)[1]
    malumotlar = ombor_malumotlarini_oqish()
    mahsulotlar = malumotlar.get(kategoriya, [])

    keyboard = types.InlineKeyboardMarkup()
    for item in mahsulotlar:
        model = item[0]
        rasm_bor = "🖼 " if (len(item) > 3 and item[3]) else ""
        keyboard.add(types.InlineKeyboardButton(
            text=f"{rasm_bor}{model}",
            callback_data=f"imgmod:{kategoriya}|{model}"
        ))

    keyboard.add(types.InlineKeyboardButton("🔙 Kategoriyalarga qaytish", callback_data="back_to_img_kat"))

    bot.send_message(
        call.message.chat.id,
        f"«<b>{kategoriya}</b>» kategoriyasidan modelni tanlang:",
        parse_mode="HTML",
        reply_markup=keyboard
    )


@bot.callback_query_handler(func=lambda call: call.data == "back_to_img_kat")
def back_to_img_kat_callback(call):
    bot.answer_callback_query(call.id)
    mahsulot_rasmi_menyu_chiqarish(call.message.chat.id, call.from_user.id)


@bot.callback_query_handler(func=lambda call: call.data == "back_to_main_admin")
def back_to_main_admin_callback(call):
    bot.answer_callback_query(call.id)
    rasm_kutilayotganlar.pop(call.from_user.id, None)
    aksiya_kutilayotganlar.pop(call.from_user.id, None)
    bot.send_message(
        call.message.chat.id,
        "🏠 Asosiy menyuga qaytdingiz.",
        reply_markup=bosh_menyu_yaratish(call.from_user.id)
    )


@bot.callback_query_handler(func=lambda call: call.data.startswith("imgmod:"))
def rasm_model_tanlandi(call):
    bot.answer_callback_query(call.id)
    if str(call.from_user.id) != str(ADMIN_ID):
        return

    kategoriya, model = call.data.split("imgmod:", 1)[1].split("|", 1)

    rasm_kutilayotganlar[call.from_user.id] = {
        "kategoriya": kategoriya,
        "model": model,
    }

    bot.send_message(
        call.message.chat.id,
        f"«<b>{model}</b>» uchun rasmni yuboring (galereyadan) yoki rasm havolasini (link) yozing:\n\n"
        f"<i>Bekor qilish uchun '🔙 Orqaga' tugmasini bosing.</i>",
        parse_mode="HTML",
        reply_markup=orqaga_menyu_yaratish()
    )


@bot.message_handler(content_types=['photo', 'text'], func=lambda message: message.from_user.id in rasm_kutilayotganlar)
def rasm_yoki_link_qabul_qilish(message):
    if message.text in ["🔙 Orqaga", "🏠 Bosh menyu"]:
        return

    holat = rasm_kutilayotganlar.get(message.from_user.id)
    if not holat:
        return

    kategoriya = holat["kategoriya"]
    model = holat["model"]

    rasm_manzili = None

    if message.photo:
        rasm_manzili = message.photo[-1].file_id
    elif message.text and message.text.startswith("http"):
        rasm_manzili = message.text.strip()
    else:
        bot.send_message(
            message.chat.id,
            "Iltimos, rasm yuboring yoki https://... bilan boshlanuvchi to'g'ri link kiriting.\n"
            "Bekor qilish uchun '🔙 Orqaga' tugmasini bosing.",
            reply_markup=orqaga_menyu_yaratish()
        )
        return

    muvaffaqiyatli = ombor_rasmini_yangilash(kategoriya, model, rasm_manzili)

    del rasm_kutilayotganlar[message.from_user.id]

    if muvaffaqiyatli:
        bot.send_message(
            message.chat.id,
            f"✅ <b>{model}</b> uchun rasm muvaffaqiyatli saqlandi!",
            parse_mode="HTML",
            reply_markup=bosh_menyu_yaratish(message.from_user.id)
        )
    else:
        bot.send_message(
            message.chat.id,
            "❌ Xatolik: ushbu model Excel'dan topilmadi.",
            reply_markup=bosh_menyu_yaratish(message.from_user.id)
        )


def aksiya_menyu_chiqarish(chat_id, user_id):
    if str(user_id) != str(ADMIN_ID):
        return

    try:
        malumotlar = ombor_malumotlarini_oqish()
    except Exception as e:
        bot.send_message(chat_id, f"Xatolik yuz berdi: {e}")
        return

    if not malumotlar:
        bot.send_message(chat_id, "Omborda mahsulot yo'q.")
        return

    keyboard = types.InlineKeyboardMarkup()
    for kategoriya in malumotlar.keys():
        keyboard.add(types.InlineKeyboardButton(
            text=f"📁 {kategoriya}",
            callback_data=f"aksiyakat:{kategoriya}"
        ))

    keyboard.add(types.InlineKeyboardButton("🔙 Orqaga", callback_data="back_to_main_admin"))

    bot.send_message(
        chat_id,
        "🎉 <b>Aksiya bo'limi:</b>\nQaysi kategoriyadagi mahsulotga aksiya biriktirasiz?",
        parse_mode="HTML",
        reply_markup=keyboard
    )
    bot.send_message(
        chat_id,
        "Ortga qaytish uchun '🔙 Orqaga' tugmasini bosing:",
        reply_markup=orqaga_menyu_yaratish()
    )


@bot.message_handler(func=lambda message: message.text == "🎉 Aksiya qo'shish")
def aksiya_boshlash(message):
    aksiya_menyu_chiqarish(message.chat.id, message.from_user.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("aksiyakat:"))
def aksiya_kategoriya_tanlandi(call):
    bot.answer_callback_query(call.id)
    if str(call.from_user.id) != str(ADMIN_ID):
        return

    kategoriya = call.data.split("aksiyakat:", 1)[1]
    malumotlar = ombor_malumotlarini_oqish()
    mahsulotlar = malumotlar.get(kategoriya, [])

    keyboard = types.InlineKeyboardMarkup()
    for item in mahsulotlar:
        model = item[0]
        aksiya_bor = "🎉 " if (len(item) > 5 and item[5]) else ""
        keyboard.add(types.InlineKeyboardButton(
            text=f"{aksiya_bor}{model}",
            callback_data=f"aksiyamod:{kategoriya}|{model}"
        ))

    keyboard.add(types.InlineKeyboardButton("🔙 Kategoriyalarga qaytish", callback_data="back_to_aksiya_kat"))

    bot.send_message(
        call.message.chat.id,
        f"«<b>{kategoriya}</b>» kategoriyasidan modelni tanlang:",
        parse_mode="HTML",
        reply_markup=keyboard
    )


@bot.callback_query_handler(func=lambda call: call.data == "back_to_aksiya_kat")
def back_to_aksiya_kat_callback(call):
    bot.answer_callback_query(call.id)
    aksiya_menyu_chiqarish(call.message.chat.id, call.from_user.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("aksiyamod:"))
def aksiya_model_tanlandi(call):
    bot.answer_callback_query(call.id)
    if str(call.from_user.id) != str(ADMIN_ID):
        return

    kategoriya, model = call.data.split("aksiyamod:", 1)[1].split("|", 1)

    aksiya_kutilayotganlar[call.from_user.id] = {
        "kategoriya": kategoriya,
        "model": model,
    }

    bot.send_message(
        call.message.chat.id,
        f"«<b>{model}</b>» uchun aksiya matnini yozing (masalan: \"20% chegirma, 10-avgustgacha\"):\n\n"
        f"<i>Aksiyani olib tashlash uchun \"-\" belgisini yuboring.\n"
        f"Bekor qilish uchun '🔙 Orqaga' tugmasini bosing.</i>",
        parse_mode="HTML",
        reply_markup=orqaga_menyu_yaratish()
    )


@bot.message_handler(func=lambda message: message.from_user.id in aksiya_kutilayotganlar)
def aksiya_matnini_qabul_qilish(message):
    if message.text in ["🔙 Orqaga", "🏠 Bosh menyu"]:
        return

    holat = aksiya_kutilayotganlar.get(message.from_user.id)
    if not holat:
        return

    kategoriya = holat["kategoriya"]
    model = holat["model"]

    matn = (message.text or "").strip()
    if not matn:
        bot.send_message(message.chat.id, "Iltimos, aksiya matnini yozing.")
        return

    aksiya_matni = "" if matn == "-" else matn
    muvaffaqiyatli = ombor_aksiyasini_yangilash(kategoriya, model, aksiya_matni)

    del aksiya_kutilayotganlar[message.from_user.id]

    if muvaffaqiyatli:
        if aksiya_matni:
            xabar = f"✅ <b>{model}</b> uchun aksiya saqlandi:\n{aksiya_matni}"
        else:
            xabar = f"✅ <b>{model}</b> uchun aksiya olib tashlandi."
        bot.send_message(
            message.chat.id, xabar, parse_mode="HTML",
            reply_markup=bosh_menyu_yaratish(message.from_user.id)
        )
    else:
        bot.send_message(
            message.chat.id,
            "❌ Xatolik: ushbu model Excel'dan topilmadi.",
            reply_markup=bosh_menyu_yaratish(message.from_user.id)
        )


@bot.callback_query_handler(func=lambda call: call.data.startswith("aksiya_korish:"))
def aksiya_korish(call):
    kategoriya, model = call.data.split("aksiya_korish:", 1)[1].split("|", 1)

    try:
        malumotlar = ombor_malumotlarini_oqish()
        mahsulotlar = {item[0]: item for item in malumotlar.get(kategoriya, [])}
        item = mahsulotlar.get(model)
        aksiya_matni = item[5] if item and len(item) > 5 and item[5] else "Bu mahsulot uchun aksiya topilmadi."
    except Exception:
        aksiya_matni = "Xatolik yuz berdi."

    bot.answer_callback_query(call.id, text=f"🎉 {model}\n\n{aksiya_matni}", show_alert=True)


@bot.message_handler(func=lambda message: message.text == "🔍 Qidiruv")
def qidiruv_boshlash(message):
    qidiruv_kutayotganlar.add(message.chat.id)
    bot.send_message(
        message.chat.id,
        "Model nomini (yoki uning bir qismini) yozing:",
        reply_markup=orqaga_menyu_yaratish()
    )


@bot.message_handler(func=lambda message: message.chat.id in qidiruv_kutayotganlar)
def qidiruv_natijasi(message):
    qidiruv_kutayotganlar.discard(message.chat.id)

    soz = (message.text or "").strip().lower()
    if not soz:
        bot.send_message(message.chat.id, "Iltimos, qidiruv uchun matn yozing.")
        return

    try:
        malumotlar = ombor_malumotlarini_oqish()
    except Exception as e:
        bot.send_message(message.chat.id, f"Xatolik yuz berdi: {e}")
        return

    topilganlar = []
    for kategoriya, mahsulotlar in malumotlar.items():
        for item in mahsulotlar:
            model, soni = item[0], item[1]
            narxi = item[2] if len(item) > 2 and item[2] else 0
            aksiya_bor = "🎉 " if (len(item) > 5 and item[5]) else ""
            if soz in model.lower() and soni and soni > 0:
                topilganlar.append((kategoriya, model, soni, narxi, aksiya_bor))

    if not topilganlar:
        bot.send_message(
            message.chat.id,
            "Hech narsa topilmadi. Boshqa so'z bilan urinib ko'ring.",
            reply_markup=bosh_menyu_yaratish(message.from_user.id)
        )
        return

    keyboard = types.InlineKeyboardMarkup()
    for kategoriya, model, soni, narxi, aksiya_bor in topilganlar[:30]:
        keyboard.add(types.InlineKeyboardButton(
            text=f"{aksiya_bor}{model} — ${narxi:,.2f} ({kategoriya})",
            callback_data=f"buy:{kategoriya}|{model}"
        ))

    bot.send_message(
        message.chat.id,
        f"Topildi: {len(topilganlar)} ta natija",
        reply_markup=keyboard
    )


def savat_matni_yaratish(chat_id):
    itemlar = savat.get(chat_id, [])
    if not itemlar:
        return "Savatingiz bo'sh.", 0

    matn = "🛒 Savatingiz:\n\n"
    jami = 0
    for item in itemlar:
        summa = item["narxi"] * item["son"]
        jami += summa
        matn += f"• {item['model']} — {item['son']} dona x ${item['narxi']:,.2f} = ${summa:,.2f}\n"

    matn += f"\n💰 Jami: ${jami:,.2f}"
    return matn, jami


def savat_tugmalari(chat_id):
    keyboard = types.InlineKeyboardMarkup()
    itemlar = savat.get(chat_id, [])

    for idx, item in enumerate(itemlar):
        keyboard.add(types.InlineKeyboardButton(
            text=f"❌ O'chirish: {item['model']}",
            callback_data=f"del_cart:{idx}"
        ))

    keyboard.add(types.InlineKeyboardButton("➕ Yana mahsulot qo'shish", callback_data="yana_mahsulot"))
    keyboard.add(types.InlineKeyboardButton("🧾 Buyurtmani rasmiylashtirish", callback_data="savat_yakunlash"))
    keyboard.add(types.InlineKeyboardButton("🗑 Savatni bo'shatish", callback_data="savat_tozalash"))
    return keyboard


@bot.message_handler(func=lambda message: message.text == "🛒 Savatim")
def menyu_savat(message):
    matn, jami = savat_matni_yaratish(message.chat.id)
    if not savat.get(message.chat.id):
        bot.send_message(message.chat.id, matn)
    else:
        bot.send_message(message.chat.id, matn, reply_markup=savat_tugmalari(message.chat.id))


@bot.callback_query_handler(func=lambda call: call.data.startswith("del_cart:"))
def savatdan_ochirish(call):
    bot.answer_callback_query(call.id)
    idx = int(call.data.split("del_cart:", 1)[1])
    chat_id = call.message.chat.id

    itemlar = savat.get(chat_id, [])
    if 0 <= idx < len(itemlar):
        ochirilgan = itemlar.pop(idx)
        bot.send_message(chat_id, f"❌ «{ochirilgan['model']}» savatdan o'chirildi.")

    matn, jami = savat_matni_yaratish(chat_id)
    if not itemlar:
        bot.send_message(chat_id, "Savatingiz bo'sh qoldi.")
    else:
        bot.send_message(chat_id, matn, reply_markup=savat_tugmalari(chat_id))


@bot.callback_query_handler(func=lambda call: call.data == "yana_mahsulot")
def yana_mahsulot_qoshish(call):
    bot.answer_callback_query(call.id)
    kategoriyalarni_korsatish(call.message.chat.id)


@bot.callback_query_handler(func=lambda call: call.data == "back_to_categories")
def back_to_categories_callback(call):
    bot.answer_callback_query(call.id)
    kategoriyalarni_korsatish(call.message.chat.id)


@bot.callback_query_handler(func=lambda call: call.data == "savat_tozalash")
def savatni_tozalash(call):
    bot.answer_callback_query(call.id)
    savat.pop(call.message.chat.id, None)
    bot.send_message(call.message.chat.id, "🗑 Savatingiz tozalandi.")


HOLAT_MATNLARI = {
    "tasdiqlangan": "✅ Tasdiqlangan",
    "bekor_qilingan": "❌ Bekor qilingan",
}


@bot.message_handler(func=lambda message: message.text == "🧾 Buyurtmalarim")
def menyu_buyurtmalarim(message):
    if str(message.from_user.id) == str(ADMIN_ID):
        buyurtmalar = barcha_buyurtmalarni_indeks_bilan_olish()

        if not buyurtmalar:
            bot.send_message(message.chat.id, "Hozircha hech qanday zakaz yo'q.")
            return

        for idx, buyurtma in buyurtmalar[:20]:
            holat_matni = HOLAT_MATNLARI.get(buyurtma.get("holat", "tasdiqlangan"), "🕓 Kutilmoqda")
            matn = f"🧾 Zakaz #{buyurtma.get('buyurtma_id', '-')}  •  {buyurtma['sana']}  •  {holat_matni}\n\n"
            for item in buyurtma["itemlar"]:
                matn += f"• {item['model']} — {item['son']} dona x ${item['narxi']:,.2f}\n"
            matn += f"\n💰 Jami: ${buyurtma['jami_summa']:,.2f}"
            matn += f"\n👤 {buyurtma['ism']}  📞 {buyurtma['telefon']}"
            matn += f"\n📍 Manzil: {buyurtma['manzil']}"

            ochirish_klaviatura = types.InlineKeyboardMarkup()
            ochirish_klaviatura.add(types.InlineKeyboardButton(
                "🗑 Tarixdan o'chirish", callback_data=f"buyurtma_ochir:{idx}"
            ))

            bot.send_message(message.chat.id, matn, reply_markup=ochirish_klaviatura)
        return

    buyurtmalar = mijoz_buyurtmalarini_olish(message.chat.id)

    if not buyurtmalar:
        bot.send_message(message.chat.id, "Sizda hali buyurtmalar yo'q.")
        return

    for buyurtma in buyurtmalar[:10]:
        holat_matni = HOLAT_MATNLARI.get(buyurtma.get("holat", "tasdiqlangan"), "🕓 Kutilmoqda")
        matn = f"🧾 {buyurtma['sana']}  •  {holat_matni}\n\n"
        for item in buyurtma["itemlar"]:
            matn += f"• {item['model']} — {item['son']} dona x ${item['narxi']:,.2f}\n"
        matn += f"\n💰 Jami: ${buyurtma['jami_summa']:,.2f}"
        matn += f"\n📍 Manzil: {buyurtma['manzil']}"
        bot.send_message(message.chat.id, matn)


@bot.message_handler(func=lambda message: message.text == "☎️ Biz bilan bog'lanish")
def menyu_kontakt(message):
    kontaktlar_matni = "\n".join(f"👤 {ism}: {tel}" for ism, tel in KONTAKTLAR)
    bot.send_message(
        message.chat.id,
        f"👥 Menedjerlar:\n\n"
        f"{kontaktlar_matni}\n\n"
        f"🕐 Ish vaqti:\n{ISH_VAQTI}"
    )


@bot.message_handler(func=lambda message: message.text == "📍 Manzilimiz")
def menyu_manzil(message):
    bot.send_message(message.chat.id, f"📍 Manzil: {KONTAKT_MANZIL}")


@bot.message_handler(func=lambda message: message.text == "ℹ️ Bot haqida")
def menyu_haqida(message):
    bot.send_message(
        message.chat.id,
        "Bu bot orqali siz ombordagi mahsulotlar qoldig'ini ko'rishingiz "
        "va zakaz berishingiz mumkin."
    )


def kategoriyalarni_korsatish(chat_id):
    try:
        malumotlar = ombor_malumotlarini_oqish()
    except FileNotFoundError:
        bot.send_message(chat_id, "Xatolik: Baza.xlsx fayli topilmadi.")
        return
    except Exception as e:
        bot.send_message(chat_id, f"Xatolik yuz berdi: {e}")
        return

    if not malumotlar:
        bot.send_message(chat_id, "Omborda hozircha mahsulot yo'q.")
        return

    keyboard = types.InlineKeyboardMarkup()
    for kategoriya in malumotlar.keys():
        keyboard.add(types.InlineKeyboardButton(
            text=kategoriya,
            callback_data=f"kat:{kategoriya}"
        ))

    bot.send_message(chat_id, "Kategoriyani tanlang:", reply_markup=keyboard)


@bot.callback_query_handler(func=lambda call: call.data.startswith("kat:"))
def kategoriya_tanlandi(call):
    bot.answer_callback_query(call.id)
    kategoriya = call.data.split("kat:", 1)[1]

    try:
        malumotlar = ombor_malumotlarini_oqish()
    except Exception as e:
        bot.send_message(call.message.chat.id, f"Xatolik yuz berdi: {e}")
        return

    mahsulotlar = malumotlar.get(kategoriya, [])
    mahsulotlar = [item for item in mahsulotlar if item[1] and item[1] > 0]

    if not mahsulotlar:
        bot.send_message(call.message.chat.id, "Bu kategoriyada hozircha omborda mahsulot yo'q.")
        return

    keyboard = types.InlineKeyboardMarkup()
    for item in mahsulotlar:
        model, soni = item[0], item[1]
        narxi = item[2] if len(item) > 2 and item[2] else 0
        aksiya_bor = "🎉 " if (len(item) > 5 and item[5]) else ""
        keyboard.add(types.InlineKeyboardButton(
            text=f"{aksiya_bor}{model} — ${narxi:,.2f}",
            callback_data=f"buy:{kategoriya}|{model}"
        ))

    keyboard.add(types.InlineKeyboardButton("🔙 Kategoriyalarga qaytish", callback_data="back_to_categories"))

    bot.send_message(
        call.message.chat.id,
        f"📦 {kategoriya}\n\nZakaz bermoqchi bo'lgan modelni tanlang:",
        reply_markup=keyboard
    )


@bot.callback_query_handler(func=lambda call: call.data.startswith("buy:"))
def zakaz_boshlash(call):
    bot.answer_callback_query(call.id)

    kategoriya, model = call.data.split("buy:", 1)[1].split("|", 1)

    malumotlar = ombor_malumotlarini_oqish()
    mahsulotlar = {
        item[0]: (item[1], item[2], item[3] if len(item) > 3 else None, item[5] if len(item) > 5 else "")
        for item in malumotlar.get(kategoriya, [])
    }
    mavjud_son, narxi, rasm, aksiya_matni = mahsulotlar.get(model, (0, 0, None, ""))

    buyurtma_holati[call.message.chat.id] = {
        "bosqich": "son",
        "kategoriya": kategoriya,
        "model": model,
        "telegram_user_id": call.from_user.id,
        "telegram_username": call.from_user.username,
    }

    matn = f"«{model}»\n💰 Narxi: ${narxi:,.2f}\n📊 Omborda: {mavjud_son} dona bor."
    if aksiya_matni:
        matn += "\n🎉 Aksiyada!"
    matn += "\n\nNechta dona kerak? Raqam bilan yozing:"

    aksiya_klaviatura = None
    if aksiya_matni:
        aksiya_klaviatura = types.InlineKeyboardMarkup()
        aksiya_klaviatura.add(types.InlineKeyboardButton(
            "🎉 Aksiya haqida", callback_data=f"aksiya_korish:{kategoriya}|{model}"
        ))

    if rasm and str(rasm).strip():
        try:
            bot.send_photo(
                call.message.chat.id,
                photo=str(rasm).strip(),
                caption=matn,
                reply_markup=orqaga_menyu_yaratish()
            )
            if aksiya_klaviatura:
                bot.send_message(call.message.chat.id, "👇", reply_markup=aksiya_klaviatura)
            return
        except Exception:
            pass

    bot.send_message(
        call.message.chat.id,
        matn,
        reply_markup=orqaga_menyu_yaratish()
    )
    if aksiya_klaviatura:
        bot.send_message(call.message.chat.id, "👇", reply_markup=aksiya_klaviatura)


@bot.callback_query_handler(func=lambda call: call.data == "savat_yakunlash")
def savatni_yakunlash_boshlash(call):
    bot.answer_callback_query(call.id)

    if not savat.get(call.message.chat.id):
        bot.send_message(call.message.chat.id, "Savatingiz bo'sh.")
        return

    buyurtma_holati[call.message.chat.id] = {
        "bosqich": "ism",
        "telegram_user_id": call.from_user.id,
        "telegram_username": call.from_user.username,
    }

    bot.send_message(
        call.message.chat.id,
        "Ismingizni yozing:",
        reply_markup=orqaga_menyu_yaratish()
    )


@bot.message_handler(func=lambda message: message.chat.id in buyurtma_holati)
def zakaz_bosqichlari(message):
    holat = buyurtma_holati.get(message.chat.id)
    if not holat:
        return

    matn = message.text.strip() if message.text else ""

    if holat["bosqich"] == "son":
        if not matn.isdigit() or int(matn) <= 0:
            bot.send_message(message.chat.id, "Iltimos, faqat musbat butun son yozing (masalan: 3).")
            return

        so_ralgan_son = int(matn)

        malumotlar = ombor_malumotlarini_oqish()
        mahsulotlar = {item[0]: (item[1], item[2]) for item in malumotlar.get(holat["kategoriya"], [])}
        mavjud_son, narxi = mahsulotlar.get(holat["model"], (0, 0))

        if so_ralgan_son > mavjud_son:
            bot.send_message(
                message.chat.id,
                f"Kechirasiz, omborda faqat {mavjud_son} dona bor. Kamroq son kiriting:"
            )
            return

        savat.setdefault(message.chat.id, []).append({
            "kategoriya": holat["kategoriya"],
            "model": holat["model"],
            "son": so_ralgan_son,
            "narxi": narxi,
        })

        del buyurtma_holati[message.chat.id]

        matn_savat, jami = savat_matni_yaratish(message.chat.id)
        bot.send_message(
            message.chat.id,
            f"✅ Savatga qo'shildi!\n\n{matn_savat}",
            reply_markup=bosh_menyu_yaratish(message.from_user.id)
        )
        bot.send_message(
            message.chat.id,
            "Tanlang:",
            reply_markup=savat_tugmalari(message.chat.id)
        )
        return

    elif holat["bosqich"] == "ism":
        if not matn:
            bot.send_message(message.chat.id, "Iltimos, ismingizni yozing:")
            return
        holat["ism"] = matn
        holat["bosqich"] = "telefon"
        bot.send_message(
            message.chat.id,
            "Telefon raqamingizni yozing (masalan: +998901234567):",
            reply_markup=orqaga_menyu_yaratish()
        )

    elif holat["bosqich"] == "telefon":
        if not matn:
            bot.send_message(message.chat.id, "Iltimos, telefon raqamingizni yozing:")
            return
        holat["telefon"] = matn
        holat["bosqich"] = "manzil"
        bot.send_message(
            message.chat.id,
            "Yetkazib berish manzilini yozing:",
            reply_markup=orqaga_menyu_yaratish()
        )

    elif holat["bosqich"] == "manzil":
        if not matn:
            bot.send_message(message.chat.id, "Iltimos, manzilni yozing:")
            return
        holat["manzil"] = matn

        itemlar = savat.get(message.chat.id, [])
        if not itemlar:
            bot.send_message(
                message.chat.id,
                "Savatingiz bo'sh qoldi, buyurtma bekor qilindi.",
                reply_markup=bosh_menyu_yaratish(message.from_user.id)
            )
            del buyurtma_holati[message.chat.id]
            return

        jami_summa = 0
        mahsulotlar_matni = ""
        for item in itemlar:
            summa = item["narxi"] * item["son"]
            jami_summa += summa
            mahsulotlar_matni += (
                f"📦 {item['model']} ({item['kategoriya']})\n"
                f"   {item['son']} dona x ${item['narxi']:,.2f} = ${summa:,.2f}\n"
            )

        buyurtma_id_hisoblagich["son"] += 1
        buyurtma_id = buyurtma_id_hisoblagich["son"]

        username_qismi = f"@{holat['telegram_username']}" if holat["telegram_username"] else "yo'q"

        kutilayotgan_buyurtmalar[buyurtma_id] = {
            "chat_id": message.chat.id,
            "itemlar": itemlar,
            "ism": holat["ism"],
            "telefon": holat["telefon"],
            "manzil": holat["manzil"],
            "jami_summa": jami_summa,
        }

        admin_xabari = (
            f"🆕 Yangi zakaz #{buyurtma_id} — tasdiq kutilmoqda\n\n"
            f"{mahsulotlar_matni}\n"
            f"💰 Jami: ${jami_summa:,.2f}\n\n"
            f"👤 Ism: {holat['ism']}\n"
            f"📞 Telefon: {holat['telefon']}\n"
            f"📍 Manzil: {holat['manzil']}\n"
            f"💬 Telegram: {username_qismi}"
        )

        tasdiq_klaviatura = types.InlineKeyboardMarkup()
        tasdiq_klaviatura.add(
            types.InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"admin_ok:{buyurtma_id}"),
            types.InlineKeyboardButton("❌ Bekor qilish", callback_data=f"admin_bekor:{buyurtma_id}"),
        )

        bot.send_message(ADMIN_ID, admin_xabari, reply_markup=tasdiq_klaviatura)

        bot.send_message(
            message.chat.id,
            "🕓 Buyurtmangiz qabul qilindi va hozir tekshirilmoqda. Tasdiqlangach sizga xabar beramiz.",
            reply_markup=bosh_menyu_yaratish(message.from_user.id)
        )

        savat.pop(message.chat.id, None)
        del buyurtma_holati[message.chat.id]


@bot.callback_query_handler(func=lambda call: call.data.startswith("admin_ok:"))
def admin_zakazni_tasdiqlash(call):
    bot.answer_callback_query(call.id)
    if str(call.from_user.id) != str(ADMIN_ID):
        return

    buyurtma_id = int(call.data.split("admin_ok:", 1)[1])
    buyurtma = kutilayotgan_buyurtmalar.pop(buyurtma_id, None)
    if not buyurtma:
        bot.send_message(call.message.chat.id, "Bu buyurtma topilmadi — avval tasdiqlangan yoki bekor qilingan bo'lishi mumkin.")
        return

    malumotlar = ombor_malumotlarini_oqish()
    ombor_matni = ""
    for item in buyurtma["itemlar"]:
        mahsulotlar_dict = {i[0]: (i[1], i[2]) for i in malumotlar.get(item["kategoriya"], [])}
        mavjud_son, _ = mahsulotlar_dict.get(item["model"], (0, 0))
        yangi_son = max(0, mavjud_son - item["son"])
        ombor_sonini_yangilash(item["kategoriya"], item["model"], yangi_son)
        ombor_matni += f"📦 {item['model']} — omborda qoldi: {yangi_son} dona\n"

    buyurtmani_saqlash(
        buyurtma["chat_id"], buyurtma["itemlar"], buyurtma["ism"],
        buyurtma["telefon"], buyurtma["manzil"], buyurtma["jami_summa"],
        holat="tasdiqlangan", buyurtma_id=buyurtma_id
    )

    bot.send_message(
        buyurtma["chat_id"],
        "✅ Zakazingiz tasdiqlandi! Tez orada siz bilan bog'lanamiz."
    )

    try:
        bot.edit_message_text(
            call.message.text + "\n\n✅ TASDIQLANDI",
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            reply_markup=None
        )
    except Exception:
        pass

    bot.send_message(
        call.message.chat.id,
        f"✅ Zakaz #{buyurtma_id} tasdiqlandi va omborda ayirildi.\n\n{ombor_matni}"
    )

    if ZAKAZ_GRUPPA_ID:
        pdf_yolu = buyurtma_pdf_yaratish(buyurtma_id, buyurtma)

        guruh_izohi = (
            f"🛒 Yangi buyurtma!\n\n"
            f"{buyurtma_raqami(buyurtma_id)}\n"
            f"Buyurtma sanasi: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
            f"🏪 Do'kon: {buyurtma['manzil']}\n"
            f"👤 Buyurtma berdi: {buyurtma['ism']}\n\n"
            f"❗️ Iltimos buyurtma bilan tanishib chiqib buyurtma statusini belgilang 👇"
        )

        status_klaviatura = types.InlineKeyboardMarkup()
        status_klaviatura.add(
            types.InlineKeyboardButton("📦 Yuk ortilmoqda", callback_data=f"status_yuk:{buyurtma_id}")
        )
        status_klaviatura.add(
            types.InlineKeyboardButton("🚚 Yo'lga chiqdi", callback_data=f"status_yolga:{buyurtma_id}")
        )

        try:
            with open(pdf_yolu, "rb") as pdf_fayl:
                bot.send_document(ZAKAZ_GRUPPA_ID, pdf_fayl, caption=guruh_izohi, reply_markup=status_klaviatura)
        except Exception:
            bot.send_message(call.message.chat.id, "⚠️ Zakaz guruhga yuborilmadi — ZAKAZ_GRUPPA_ID to'g'ri sozlanganini tekshiring.")
        finally:
            if os.path.exists(pdf_yolu):
                os.remove(pdf_yolu)


@bot.callback_query_handler(func=lambda call: call.data.startswith("status_yuk:"))
def status_yuk_belgilash(call):
    bot.answer_callback_query(call.id, "Belgilandi!")

    buyurtma_id = call.data.split("status_yuk:", 1)[1]
    yangi_izoh = call.message.caption + "\n\n✅ 📦 Yuk ortilmoqda - statusi belgilandi."

    yangi_klaviatura = types.InlineKeyboardMarkup()
    yangi_klaviatura.add(
        types.InlineKeyboardButton("🚚 Yo'lga chiqdi", callback_data=f"status_yolga:{buyurtma_id}")
    )

    try:
        bot.edit_message_caption(
            caption=yangi_izoh,
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            reply_markup=yangi_klaviatura
        )
    except Exception:
        pass


@bot.callback_query_handler(func=lambda call: call.data.startswith("status_yolga:"))
def status_yolga_belgilash(call):
    bot.answer_callback_query(call.id, "Belgilandi!")

    yangi_izoh = call.message.caption + "\n\n✅ 🚚 Yo'lga chiqdi — statusi belgilandi!"

    try:
        bot.edit_message_caption(
            caption=yangi_izoh,
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            reply_markup=None
        )
    except Exception:
        pass


@bot.callback_query_handler(func=lambda call: call.data.startswith("admin_bekor:"))
def admin_zakazni_bekor_qilish(call):
    bot.answer_callback_query(call.id)
    if str(call.from_user.id) != str(ADMIN_ID):
        return

    buyurtma_id = int(call.data.split("admin_bekor:", 1)[1])
    buyurtma = kutilayotgan_buyurtmalar.pop(buyurtma_id, None)
    if not buyurtma:
        bot.send_message(call.message.chat.id, "Bu buyurtma topilmadi — avval tasdiqlangan yoki bekor qilingan bo'lishi mumkin.")
        return

    buyurtmani_saqlash(
        buyurtma["chat_id"], buyurtma["itemlar"], buyurtma["ism"],
        buyurtma["telefon"], buyurtma["manzil"], buyurtma["jami_summa"],
        holat="bekor_qilingan", buyurtma_id=buyurtma_id
    )

    bot.send_message(
        buyurtma["chat_id"],
        "❌ Afsuski, buyurtmangiz bekor qilindi. Batafsil ma'lumot uchun biz bilan bog'laning."
    )

    try:
        bot.edit_message_text(
            call.message.text + "\n\n❌ BEKOR QILINDI",
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            reply_markup=None
        )
    except Exception:
        pass

    bot.send_message(call.message.chat.id, f"❌ Zakaz #{buyurtma_id} bekor qilindi. Ombor o'zgartirilmadi.")


@bot.callback_query_handler(func=lambda call: call.data.startswith("buyurtma_ochir:"))
def buyurtmani_tarixdan_ochirish(call):
    bot.answer_callback_query(call.id)
    if str(call.from_user.id) != str(ADMIN_ID):
        return

    idx = int(call.data.split("buyurtma_ochir:", 1)[1])

    try:
        with open(BUYURTMALAR_FAYLI, "r", encoding="utf-8") as f:
            barcha = json.load(f)
    except Exception:
        barcha = []

    if 0 <= idx < len(barcha):
        ochirilgan = barcha.pop(idx)

        tiklangan_matn = ""
        if ochirilgan.get("holat", "tasdiqlangan") == "tasdiqlangan":
            try:
                malumotlar = ombor_malumotlarini_oqish()
            except Exception:
                malumotlar = {}

            for item in ochirilgan.get("itemlar", []):
                mahsulotlar_dict = {i[0]: i[1] for i in malumotlar.get(item["kategoriya"], [])}
                hozirgi_son = mahsulotlar_dict.get(item["model"], 0) or 0
                yangi_son = hozirgi_son + item["son"]
                ombor_sonini_yangilash(item["kategoriya"], item["model"], yangi_son)
                tiklangan_matn += f"📦 {item['model']} — omborga qaytarildi: +{item['son']} (endi {yangi_son} dona)\n"

        with open(BUYURTMALAR_FAYLI, "w", encoding="utf-8") as f:
            json.dump(barcha, f, ensure_ascii=False, indent=2)

        xabar = call.message.text + "\n\n🗑 TARIXDAN O'CHIRILDI"
        if tiklangan_matn:
            xabar += f"\n\n🔄 Ombor tiklandi:\n{tiklangan_matn}"

        try:
            bot.edit_message_text(
                xabar,
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                reply_markup=None
            )
        except Exception:
            pass
    else:
        bot.send_message(call.message.chat.id, "Bu yozuv topilmadi — allaqachon o'chirilgan bo'lishi mumkin.")


@bot.message_handler(content_types=['document'])
def yangi_fayl_qabul_qilish(message):
    if str(message.from_user.id) != str(ADMIN_ID):
        return

    fayl_nomi = message.document.file_name or ""
    if not fayl_nomi.lower().endswith(".xlsx"):
        bot.send_message(message.chat.id, "Faqat .xlsx fayl yuboring.")
        return

    try:
        fayl_info = bot.get_file(message.document.file_id)
        yuklab_olingan = bot.download_file(fayl_info.file_path)

        with open(EXCEL_FILE, "wb") as f:
            f.write(yuklab_olingan)

        malumotlar = ombor_malumotlarini_oqish()
        jami_mahsulot = sum(len(v) for v in malumotlar.values())

        bot.send_message(
            message.chat.id,
            f"✅ Ombor fayli yangilandi!\n"
            f"Jami {len(malumotlar)} ta kategoriya, {jami_mahsulot} ta model topildi."
        )
    except Exception as e:
        bot.send_message(message.chat.id, f"Xatolik yuz berdi: {e}")


@bot.message_handler(commands=['yangilash'])
def yangilash_boshlash(message):
    if str(message.from_user.id) != str(ADMIN_ID):
        bot.send_message(message.chat.id, "Kechirasiz, bu buyruq faqat admin uchun.")
        return

    try:
        malumotlar = ombor_malumotlarini_oqish()
    except Exception as e:
        bot.send_message(message.chat.id, f"Xatolik yuz berdi: {e}")
        return

    if not malumotlar:
        bot.send_message(message.chat.id, "Omborda hozircha mahsulot yo'q.")
        return

    keyboard = types.InlineKeyboardMarkup()
    for kategoriya in malumotlar.keys():
        keyboard.add(types.InlineKeyboardButton(
            text=kategoriya,
            callback_data=f"yangkat:{kategoriya}"
        ))

    bot.send_message(message.chat.id, "Qaysi kategoriyadagi mahsulotni yangilaysiz?", reply_markup=keyboard)


@bot.callback_query_handler(func=lambda call: call.data.startswith("yangkat:"))
def yangilash_kategoriya_tanlandi(call):
    bot.answer_callback_query(call.id)

    if str(call.from_user.id) != str(ADMIN_ID):
        return

    kategoriya = call.data.split("yangkat:", 1)[1]
    malumotlar = ombor_malumotlarini_oqish()
    mahsulotlar = malumotlar.get(kategoriya, [])

    if not mahsulotlar:
        bot.send_message(call.message.chat.id, "Bu kategoriyada mahsulot topilmadi.")
        return

    keyboard = types.InlineKeyboardMarkup()
    for item in mahsulotlar:
        model, soni = item[0], item[1]
        keyboard.add(types.InlineKeyboardButton(
            text=f"{model} (hozir: {soni} dona)",
            callback_data=f"yangmodel:{kategoriya}|{model}"
        ))

    bot.send_message(call.message.chat.id, "Qaysi modelni yangilaysiz?", reply_markup=keyboard)


@bot.callback_query_handler(func=lambda call: call.data.startswith("yangmodel:"))
def yangilash_model_tanlandi(call):
    bot.answer_callback_query(call.id)

    if str(call.from_user.id) != str(ADMIN_ID):
        return

    kategoriya, model = call.data.split("yangmodel:", 1)[1].split("|", 1)

    yangilash_holati[call.from_user.id] = {
        "kategoriya": kategoriya,
        "model": model,
    }

    bot.send_message(
        call.message.chat.id,
        f"«{model}» uchun yangi sonni raqam bilan yuboring (masalan: 15):",
        reply_markup=orqaga_menyu_yaratish()
    )


@bot.message_handler(func=lambda message: message.from_user.id in yangilash_holati)
def yangi_sonni_qabul_qilish(message):
    holat = yangilash_holati.get(message.from_user.id)
    if not holat:
        return

    matn = message.text.strip()
    if not matn.isdigit():
        bot.send_message(message.chat.id, "Iltimos, faqat butun son yuboring (masalan: 15).")
        return

    yangi_soni = int(matn)
    kategoriya = holat["kategoriya"]
    model = holat["model"]

    muvaffaqiyatli = ombor_sonini_yangilash(kategoriya, model, yangi_soni)

    del yangilash_holati[message.from_user.id]

    if muvaffaqiyatli:
        bot.send_message(
            message.chat.id,
            f"✅ Yangilandi: {model} — endi {yangi_soni} dona.",
            reply_markup=bosh_menyu_yaratish(message.from_user.id)
        )
    else:
        bot.send_message(
            message.chat.id,
            "Xatolik: bu model Excel'da topilmadi.",
            reply_markup=bosh_menyu_yaratish(message.from_user.id)
        )


print("Bot ishga tushdi...")
bot.infinity_polling()
